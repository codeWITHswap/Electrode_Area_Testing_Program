# import the modules
import os
from os import listdir
import cv2
import numpy as np
from openpyxl import *

def empty(a):
        pass

def process_test_img(img_name, img, roi, imgContour, scaling_factor, minDist_usr, param1_usr, param2_usr, thresh1_usr, thresh2_usr, minArea_usr):
        print(img_name)
        gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        # blur the grayscale image
        # gray_blurred = cv2.medianBlur(gray,5)
        # We will use cv2.HoughCircles() function to detect circles in a grayscale image (It does so by using Hough Transform)

        circles = cv2.HoughCircles(gray, cv2.HOUGH_GRADIENT, 1, minDist_usr, param1_usr, param2_usr, minRadius=0, maxRadius=0)
        detected_circles = np.uint16(np.around(circles))

        list_r = []
        for i in detected_circles[0, :]:
                list_r.append(i[2])
        X = 0
        Y = 0
        R = 0
        for i in detected_circles[0, :]:
                r = i[2]
                if r == max(list_r):
                        X = i[0]
                        Y = i[1]
                        R = r
                        roi = img[Y-R:Y+R, X-R:X+R]

        # imgBlur = cv2.GaussianBlur(roi, (3,3), 1) # Step 1: Blur the image
        imgGray = cv2.cvtColor(roi,cv2.COLOR_BGR2GRAY) # Step 2: Convert the blurred image into Black&White image

        imgCanny = cv2.Canny(imgGray,thresh1_usr,thresh2_usr)
        kernel = np.ones((2,2))
        imgDil = cv2.dilate(imgCanny, kernel, iterations = 1)

        ret, thresh1 = cv2.threshold(imgGray, 137, 255, cv2.THRESH_BINARY)
        ret, thresh2 = cv2.threshold(imgGray, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)

        contours, hierarchy = cv2.findContours(thresh2, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
        for i in range(len(contours)):
                for j in range(len(contours[i])):
                        for k in range(len(contours[i][j])):
                                contours[i][j][k][0]+=(X-R)
                                contours[i][j][k][1]+=(Y-R)

        contour_areas=[]
        print(img_name)
        for cnt in contours:
            area = cv2.contourArea(cnt)
            if area > minArea_usr:
                cv2.drawContours(imgContour, cnt, -1, (0,0,255),1)
                contour_areas.append(area)
                print(area)
        print("\n")

        size = len(img_name)
        contour_name = img_name[:size-4]+"_contour.jpg"
        cv2.imwrite(contour_name,imgContour)

        # contour_areas.sort()

        # current_row = sheet.max_row
        # sheet.cell(row=current_row+1, column=1).value = img_name
        # sheet.cell(row=current_row+1, column=2).value = contour_areas[-1]*scaling_factor
        # sheet.cell(row=current_row+1, column=3).value = contour_areas[-2]*scaling_factor
        # sheet.cell(row=current_row+1, column=4).value = contour_areas[-1]
        # sheet.cell(row=current_row+1, column=5).value = contour_areas[-2]
        # wb.save(filename = 'Results.xlsx')
                
        # size = len(img_name)
        # contour_name = img_name[:size-4]+"_contour.jpg"
        # cv2.imwrite(contour_name,imgContour)


# get the path to the folder
path1 = input("Path to Image Folder: ")
ref_area = int(input("Reference Area: "))

# cv2.namedWindow("Parameters")
# cv2.resizeWindow("Parameters",100,100)
# cv2.createTrackbar("Threshold1","Parameters",255,255,empty)  
# cv2.createTrackbar("Threshold2","Parameters",255,255,empty)
# cv2.createTrackbar("minArea","Parameters",1047,50000,empty)
# cv2.createTrackbar("minDist","Parameters",99,1000,empty)
# cv2.createTrackbar("Param1","Parameters",15,100,empty)
# cv2.createTrackbar("Param2","Parameters",18,100,empty)

os.chdir(path1)

# wb = Workbook()
# sheet =  wb.active
# sheet.title = "Electrode Areas"

# sheet.cell(row=1, column=1).value = "Image Name"
# sheet.cell(row=1, column=2).value = "Relative Working Electrode Area"
# sheet.cell(row=1, column=3).value = "Relative Counter Electrode Area"
# sheet.cell(row=1, column=4).value = "True Working Electrode Area"
# sheet.cell(row=1, column=5).value = "True Counter Electrode Area"
# wb.save(filename = 'Results.xlsx')



# change the directory to the given path

list=[]
for images in os.listdir(path1):

        # check if the image ends with png or jpg or jpeg
        if (images.endswith(".png") or images.endswith(".jpg") or images.endswith(".jpeg")):
                list.append(images)

cv2.namedWindow("Parameters")
cv2.resizeWindow("Parameters",500,500)
cv2.createTrackbar("minDist","Parameters",124,1000,empty)
cv2.createTrackbar("Param1","Parameters",40,100,empty)
cv2.createTrackbar("Param2","Parameters",38,100,empty)

while True : 
    img = cv2.imread("ref.png")

    circled_img = img.copy()
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    # gray_blurred = cv2.medianBlur(gray,5)
    
    if cv2.getTrackbarPos("minDist","Parameters") > 1:
        minDistance = cv2.getTrackbarPos("minDist","Parameters")
    else:
        minDistance = 1

    if cv2.getTrackbarPos("Param1","Parameters") > 10:
        param1 = cv2.getTrackbarPos("Param1","Parameters")
    else:
        param1 = 10

    if cv2.getTrackbarPos("Param2","Parameters") > 10:
        param2 = cv2.getTrackbarPos("Param2","Parameters")
    else:
        param2 = 10


    circles = cv2.HoughCircles(gray, cv2.HOUGH_GRADIENT, 1, minDistance , param1, param2, minRadius=0, maxRadius=0)
    if circles is not None:
    # Convert the circle parameters a, b and r to integers.
        detected_circles = np.uint16(np.around(circles))
        for pt in detected_circles[0, :]:
            a, b, r = pt[0], pt[1], pt[2]
            # Draw the circumference of the circle.
            cv2.circle(circled_img, (a, b), r, (0, 255, 0), 2)
        cv2.imshow("Detected Circle", circled_img)
        
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

minDist_usr=cv2.getTrackbarPos("minDist","Parameters")
param1_usr=cv2.getTrackbarPos("Param1","Parameters")
param2_usr=cv2.getTrackbarPos("Param2","Parameters")
print(minDist_usr, param1_usr, param2_usr)
cv2.destroyWindow("Parameters")
list_r = []
for i in detected_circles[0, :]:
    list_r.append(i[2])
X = 0
Y = 0
R = 0
for i in detected_circles[0, :]:
        r = i[2]
        if r == max(list_r):
            X = i[0]
            Y = i[1]
            R = r
            roi = img[Y-R:Y+R, X-R:X+R]


cv2.namedWindow("Parameters")
cv2.resizeWindow("Parameters",500,500)
cv2.createTrackbar("Threshold1","Parameters",255,255,empty)  
cv2.createTrackbar("Threshold2","Parameters",255,255,empty)
cv2.createTrackbar("minArea","Parameters",5497,50000,empty)


while True:
    imgContour = cv2.imread("ref.png")
    # imgBlur = cv2.GaussianBlur(roi, (3,3), 1)
    imgGray = cv2.cvtColor(roi,cv2.COLOR_BGR2GRAY) # Step 2: Convert the blurred image into Black&White image


    if cv2.getTrackbarPos("Threshold1","Parameters") > 1:
        thresh1 = cv2.getTrackbarPos("Threshold1","Parameters")
    else:
        thresh1 = 1

    if cv2.getTrackbarPos("Threshold2","Parameters") > 1:
        thresh2 = cv2.getTrackbarPos("Threshold2","Parameters")
    else:
        thresh2 = 1

    if cv2.getTrackbarPos("minArea","Parameters") > 1:
        minArea = cv2.getTrackbarPos("minArea","Parameters")
    else:
        minArea = 1


    imgCanny = cv2.Canny(imgGray,thresh1,thresh2)
    kernel = np.ones((2,2)) # To reduce the dilation, try using smaller kernels like (2,2)
    imgDil = cv2.dilate(imgCanny, kernel, iterations = 1)

    ret, thresh1 = cv2.threshold(imgGray, 137, 255, cv2.THRESH_BINARY)
    ret, thresh2 = cv2.threshold(imgGray, 0, 255, cv2.THRESH_BINARY+cv2.THRESH_OTSU)

    contours, hierarchy = cv2.findContours(thresh2, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
    for i in range(len(contours)):
        for j in range(len(contours[i])):
            for k in range(len(contours[i][j])):
                contours[i][j][k][0]+=(X-R)
                contours[i][j][k][1]+=(Y-R)

    for cnt in contours:
        area = cv2.contourArea(cnt)
        if area > minArea:
            cv2.drawContours(imgContour, cnt, -1, (0,0,255),1)
    cv2.imshow("Detected Circle", imgContour)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break
            
thresh1_usr=cv2.getTrackbarPos("Threshold1","Parameters")
thresh2_usr=cv2.getTrackbarPos("Threshold2","Parameters")
minArea_usr=cv2.getTrackbarPos("minArea","Parameters")
print(thresh1_usr, thresh2_usr, minArea_usr)
cv2.destroyWindow("Parameters")

list_area = []
print("ref.png")
for cnt in contours:
        area = cv2.contourArea(cnt)
        list_area.append(area)
        print(area, "\n")
list_area.sort() 
scaling_factor = ref_area / list_area[-1]

# current_row = sheet.max_row
# sheet.cell(row=current_row+1, column=1).value = "ref.png"
# sheet.cell(row=current_row+1, column=2).value = list_area[-1]*scaling_factor
# sheet.cell(row=current_row+1, column=3).value = list_area[-2]*scaling_factor
# sheet.cell(row=current_row+1, column=4).value = list_area[-1]
# sheet.cell(row=current_row+1, column=5).value = list_area[-2]
# wb.save(filename = 'Results.xlsx')

img_name = "ref.png"
size = len(img_name)
contour_name = img_name[:size-4]+"_contour.jpg"
cv2.imwrite(contour_name,imgContour)

for img_name in range(len(list)):
    if not list[img_name] == "ref.jpg" and not list[img_name] == "ref.png":

        orig_img = cv2.imread(list[img_name])
        roi = orig_img.copy()
        imgContour = orig_img.copy()
        process_test_img(list[img_name],orig_img,roi,imgContour, scaling_factor, minDist_usr, param1_usr, param2_usr, thresh1_usr, thresh2_usr, minArea_usr)
